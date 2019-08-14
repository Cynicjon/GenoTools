#!/usr/bin/env python3
import configparser
import os
from sys import argv
import re

import PIL
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule

from Monitor import Message


class Export(object):

    def __init__(self):
        """Initialise the exporter. Loading files here means they only need to be loaded once."""
        self.cols_order = ['Well ', 'Omitted ', 'Sample', 'Target', 'Reporter', 'RQ   ', 'Cт', 'ΔCт', 'ΔΔCт', 'Mouse',
                           'Genotype', 'Allele', 'Locked', 'Plate Barcode', 'Assay Type', 'Assay Name', 'Result',
                           'Confirmed', 'Comment', 'Name', 'Compare', 'Gender', 'Het Control?', 'X-Linked?',
                           'Omitted_endo']
        self.config = configparser.ConfigParser()

        if os.path.isfile(os.getcwd() + '/config.ini'):  # may have changed.
            self.config.read(os.getcwd() + '/config.ini')
        else:
            self.config.read(os.path.normpath(os.path.dirname(argv[0]) + '/config.ini'))

        self.assay_df = self.read_assay_file()  # Reads Assay info from file
        self.genf, self.assayf, self.confirmf = self.read_formulas()

        self.inp = self.samples = self.ctrls = self.ctrl_targets = self.ctrl_name = self.endo = None

        self.xlsx_file = None           # Str: path
        self._last_file = None          # Str: path
        self._multi_export = False      # Bool

    def new(self, inp: str):
        """
        This is called by Monitor.Labhandler.On_Created(), and takes the input from csv through to completed file.
        :param inp: file path of exported csv.
        """
        self.inp = inp
        self.samples = self.read_file()
        self.read_endo_ctrl()
        self.endo_cleanup()
        self.separate_ctrls()
        self.samples['Assay Type'] = self.samples.apply(lambda line: pd.Series([self.assay_type(line)]), axis=1)
        self.samples = self.samples.sort_values(by=['Assay Type', 'Target', 'Sample'])  # Sort rows
        self.add_formulas()
        # Insert ctrls to end of file, sort + remove unneeded columns
        self.samples = pd.concat([self.samples, self.ctrls], sort=True)[self.cols_order]
        # CT floats are str because of a mixture of strings and floats, this causes 'number formatted as text' flags in
        # excel. Here we convert back into floats where possible.
        self.samples['Cт'] = pd.to_numeric(self.samples['Cт'], errors='coerce')
        self.samples['Cт'] = self.samples['Cт'].fillna("Undetermined")
        try:
            self.to_xlsx()
        except PermissionError as e:
            print(e)
            print(Message("You already have an export of this file open. Close it and re-try.").red())
        except ValueError as e:  # if file is missing cols or is not an export - raised in read_file()
            print(e)

    def read_file(self):
        """Reads the file given as input and returns a dataframe. Only accepts columns in the first 9 of cols_order."""
        params = [(14, "\t"), (15, "\t"), (14, ","), (15, ",")]  # list of parameters to try
        for header, sep in params:
            try:
                return pd.read_csv(self.inp, header=header, sep=sep, usecols=self.cols_order[:9])
            except ValueError:
                continue
        else:
            raise ValueError(
                "That file doesnt look right.\nIt is either missing a column we need, or it is not an export" +
                " file.\nCheck that you ticked 'Results' when exporting.\nColumns needed: " +
                ", ".join(self.cols_order[:9]) + ".")

    def read_endo_ctrl(self):
        """
        Reads CSV to get control name, and endogenous control name, then uses control name to get a list of targets
        that the control applies to.
        """
        with open(self.inp, "r") as file:
            for i in range(4):
                next(file)
            self.endo = file.readline().split(sep='=')[1].strip().lower()
            for i in range(7):
                next(file)
            self.ctrl_name = file.readline().split(sep='=')[1].strip().lower()
        self.ctrl_targets = set(self.samples[self.samples.Sample.str.lower() == self.ctrl_name]['Target'].tolist())

    def endo_cleanup(self):
        """
        Removes endogenous control results. Adds a field to samples that indicates if the corresponding endo was omitted
        """
        endos = self.samples[self.samples.Target.str.lower() == self.endo][['Well ', 'Omitted ']]  # get needed cols
        endos = endos.rename(index=str, columns={'Omitted ': 'Omitted_endo'})  # rename column
        self.samples = self.samples[self.samples.Target.str.lower() != self.endo]  # Remove endo from samples
        self.samples = pd.merge(self.samples, endos, on='Well ', how='inner')  # Merge endos with samples

    def read_formulas(self):
        """Loads excel formulas used from a file, calls get_formula_sub to format them and sets the finished formulas"""
        try:
            wb = load_workbook(self.config['File paths']['Formulas'])
        except FileNotFoundError:
            wb = input("Can't find the Formulas file! If it has moved, please update config.ini with its new location."
                       "\nPress Enter to quit")
            quit()
        return self.get_formula_sub(wb['Sheet1'].cell(row=2, column=11).value),\
            self.get_formula_sub(wb['Sheet1'].cell(row=2, column=17).value),\
            self.get_formula_sub(wb['Sheet1'].cell(row=2, column=18).value)

    @staticmethod
    def get_formula_sub(formula):
        """Takes in a formula, splits it and replaces the row number with a sub string {0} for later use."""
        from openpyxl.formula import Tokenizer
        formula_sub = "="
        for t in Tokenizer(formula).items:  # Tokenizer is part of openpyxl
            if t.subtype is 'RANGE':  # if token is a cell reference
                t.value = t.value[0] + "{0}"  # replace row number with {0} e.g F3 -> F{0}
                formula_sub += t.value  # add it to formula_sub
            else:
                formula_sub += t.value  # else do nothing and add it to formula_sub
        return formula_sub

    def read_assay_file(self):
        """Reads the assay file and returns a dataframe. The assay file path is specified in config.ini"""
        try:
            assays = pd.read_csv(self.config['File paths']['assays'], sep='\t')  # reads file.
            assays['Variant'] = assays['Variant'].str.lower()  # sets variant col to lower-case
            return assays.set_index('Variant', drop=False)  # set variant col as index, keeping variant as a column
        except FileNotFoundError:
            input("Can't find the Assays file! If it has moved, please update config.ini with its new location."
                  "\nPress Enter to quit.")
            quit()

    def separate_ctrls(self):
        """ Move controls out of the samples df and into a ctrls df, and sort.
        Regular expression pattern matches a mouse or blastocyst in the form ~PMGB11.2a or M02983000"""
        regex = '[a-z]{3,4}\\d{1,3}\\.\\d{1,2}[a-z]|m\\d{8}'
        self.ctrls = self.samples.loc[~self.samples['Sample'].str.lower().str.contains(regex)]  # Move ctrls to new df
        self.samples = self.samples.loc[self.samples['Sample'].str.lower().str.contains(regex)]  # Remove ctrls from df
        self.ctrls = self.ctrls.sort_values(by=['Target', 'Sample'])

    def assay_type(self, line):
        """Determines if the assay type is LoA or qPCR. List of assays is read in from a file."""
        target = line['Target'].lower()
        if target in self.assay_df['Variant']:      # If the target is in list of assays, set assay type accordingly.
            return self.assay_df.loc[target, 'Type']
        elif "_wt" in target or "_ce" in target:
            return "LoA"
        else:
            return "Unknown"

    def add_formulas(self):
        """ Adds formulas and extra columns. Row number is added by string formatting based on df['index']
            Adding formulas rather than doing the logic in python allows the user to make adjustments."""

        self.samples['index'] = range(2, self.samples.shape[0] + 2)  # make index == to excel row number
        barcode = os.path.basename(self.inp).split("_")[0].upper()     # get plate barcode from input file path
        self.samples['Assay Name'] = self.samples.apply(lambda line: pd.Series([self.assay_name(line)]), axis=1)
        columns_add = {'Mouse': self.samples['Sample'], 'Plate Barcode': barcode,
                       'Allele': np.nan, 'Locked': np.nan, 'Comment': np.nan, 'Name': np.nan,
                       'Compare': np.nan, 'Gender': np.nan,

                       'Het Control?': self.samples.apply(lambda line: pd.Series([self.is_het_ctrl(line)]), axis=1),
                       'X-Linked?': self.samples.apply(lambda line: pd.Series([self.is_transgene(line)]), axis=1),
                       'RQ   ': self.samples.apply(lambda line: pd.Series([self.rq_add_zero(line)]), axis=1),

                       'Genotype': self.samples.apply(lambda line: pd.Series([self.genf.format(line['index'])]),
                                                      axis=1),
                       'Result': self.samples.apply(lambda line: pd.Series([self.assayf.format(line['index'])]),
                                                    axis=1),
                       'Confirmed': self.samples.apply(lambda line: pd.Series([self.confirmf.format(line['index'])]),
                                                       axis=1)
                       }  # dict of columns we need to add and their values
        for col_name in columns_add:  # Add the columns in columns_add, and set its value respectively.
            self.samples[col_name] = columns_add[col_name]

    def assay_name(self, line):
        """
        Checks Target against a list of common spelling mistakes and corrects them, for uniformity in the database.
        """
        target = line['Target'].lower()
        if line['Target'].lower() in self.assay_df['Variant']:
            return self.assay_df.loc[target, 'Assay']
        else:
            return line['Target']

    def is_het_ctrl(self, line):
        """
        If the control in export is a het, and applies to that target, return "Yes". The excel formula will adjust
        the analysis accordingly.
        """
        if "het" in self.ctrl_name and line['Target'] in self.ctrl_targets:
            return "Yes"

    @staticmethod
    def is_transgene(line):
        """If the assay is a transgene assay, returns Transgene. The excel formula adjusts the analysis accordingly."""
        target = line['Assay Name'].upper()
        if "_TG" in target:
            return "Transgene"

    def rq_add_zero(self, line):
        """
        If the Cт value is undetermined, the RQ value should be 0 rather than None. This should only be done if the
        control applies to that sample/target and should not be done if the endogenous control has been omitted.
        """
        if line['Omitted_endo']:
            return np.nan  # If endo is omitted, insert NaN.
        if line['Cт'] == "Undetermined":  # If CT is Undetermined
            # Add 0 to RQ if target also applies to the control
            return 0 if line['Target'] in self.ctrl_targets else np.nan
        else:
            return line['RQ   ']  # Keep same RQ value.

    @property
    def multi(self):
        """
        Returns string indicating if Multi export mode is on of off. If multi export is OFF and there is an xlsz file,
        then multi must have been toggled off recently, and the file is launched etc.
        """
        if not self._multi_export and self.xlsx_file:
            os.startfile(self.xlsx_file)  # Try/except shouldn't be needed here.
            print(Message(' ' + os.path.split(self.xlsx_file)[1]).timestamp('Export'))
            self._last_file = self.xlsx_file
            self.xlsx_file = None
        return Message(''.ljust(25, ' ') + 'Multi export processing ON') if self._multi_export \
            else Message(''.ljust(25, ' ') + 'Multi export processing OFF')

    @multi.setter
    def multi(self, value):
        # Sets _multi_export flag with value, if value is none, toggles _multi_export. Prints multi.
        if value is not None:
            if not self._multi_export == value:  # Only change value (and print) if value changes
                self._multi_export = value
                print(self.multi)
        else:
            self._multi_export = not self._multi_export
            print(self.multi)

    def multi_toggle(self):
        self.multi = None

    def multi_off(self):
        self.multi = False

    def last_file(self):
        self.xlsx_file = self._last_file
        print(''.ljust(25, ' ') + "Exporting to last exported file.")

    def to_file(self):
        # Allows Input of a specific xlsx file to export to.
        from Monitor import InputLoop
        print(''.ljust(25, ' ') + 'Enter a target file (.xlsx) or type stop to cancel')
        while True:
            inp = InputLoop.get_input()
            if os.path.isfile(inp) and inp[-5:] == '.xlsx':
                self.xlsx_file = inp
                break
            if inp.lower() == 'stop':
                self.multi = False
                break
            else:
                print(Message('That isn\'t an excel file path!').red())
        if self.xlsx_file:
            print(''.ljust(25, ' ') + 'Thanks. You can now export your files, or paste the file path here.')

    def get_sheet_name(self):
        """Parses the file name and shortens it to <32 chars so it can be used as the sheet name in excel."""
        plate = str(os.path.split(os.path.splitext(self.inp)[0])[1])  # unsure why or if str() needed, pycharm likes it
        plate_barcode = r'^c0000\d{5}' + r'|^sl000\d{5}' + r'|^\d{5}'
        rex = re.compile(plate_barcode, re.IGNORECASE)
        plate2 = plate.split(sep='_')
        try:
            plate2.remove('data')
        except ValueError:
            pass
        # It is difficult to separate user names from gene names like cd4 etc, so we use a list of user names.
        users = self.config['Users']['users'].split(',')
        users.append(os.getlogin())
        user, plates, assays_etc, plates_small = [], [], [], []  # 4 lists representing what the filename is split into
        for item in plate2:
            if item in users:  # If the element is a username add to user list
                user.append(item)
                continue
            match = rex.match(item)  # If the item is a barcode, add to plates list
            plates.append(match.group()) if match else assays_etc.append(item)

        for i in plates:  # Make a list of shortened plate names.
            plates_small.append(i.upper().replace('SL000', '').replace('C0000', '')[:5])

        final = '_'.join(plates + assays_etc + user)
        # There is a max sheet name length of 32 chars in excel. Try shortening steps whilst keeping info if possible
        if len(final) >= 31:    # There is a max sheet name length of 32 chars in excel.
            final = '_'.join(plates_small + assays_etc + user)  # remove SL000 or C0000
        if len(final) >= 31:
            final = '_'.join(plates_small[0:1] + assays_etc + user)  # Only use main plate barcode
        if len(final) >= 31:
            final = '_'.join(plates_small[0:1] + assays_etc)  # drop the username
        if len(final) >= 31:
            final = final[:31]  # Truncate.
        return final

    def to_xlsx(self):
        """
        Exports to xlsx file and formats it with correct column widths, top row freeze pane and conditional formatting
        based on genotype.
        """
        sheet = self.get_sheet_name()
        if not self.xlsx_file:
            self.xlsx_file = os.path.splitext(self.inp)[0] + '.xlsx'

        if os.path.isfile(self.xlsx_file):
            self.multi = True
            book = load_workbook(self.xlsx_file)
            writer = pd.ExcelWriter(self.xlsx_file, engine='openpyxl')
            writer.book = book
            if sheet in book.sheetnames:  # if the sheet already exists, add a digit on the end.
                for i in range(1, 100):
                    sheet1 = sheet + str(i)
                    if len(sheet1) > 31:
                        sheet1 = sheet[:30] + str(i)
                    if len(sheet1) > 31:
                        sheet1 = sheet[:29] + str(i)
                    if sheet1 not in book.sheetnames:
                        sheet = sheet1
                        break
        else:
            writer = pd.ExcelWriter(self.xlsx_file, engine='openpyxl')
        self.samples.to_excel(writer, sheet_name=sheet, index=False, freeze_panes=(1, 0))  # Write dataframe to excel
        """Formatting for a pretty output"""
        wb = writer.book
        ws = wb[sheet]

        col_width = {'A': 5, 'C': 11, 'J': 11, 'K': 9.14, 'N': 12.57, 'O': 10, 'P': 18, 'R': 10, 'S': 15.14, 'W': 11.43,
                     'Y': 13.43}
        for col in col_width:  # set column widths
            ws.column_dimensions[col].width = col_width[col]
            for cell in ws[col]:    # set center align
                cell.alignment = Alignment(horizontal='center')

        conditions = {'Het': PatternFill(patternType='solid', bgColor='DCE6F0'),
                      'Hom': PatternFill(patternType='solid', bgColor='B8CCE4'),
                      'Hemi': PatternFill(patternType='lightUp', bgColor='DCE6F0', fgColor='B8CCE4'),
                      'Fail': PatternFill(patternType='solid', bgColor='FFC7CE'),
                      'Retest': PatternFill(patternType='solid', bgColor='FFC7CE')}
        for genotype in conditions:  # Add conditional formatting
            ws.conditional_formatting.add('K2:K' + str(self.samples.shape[0] + 2),
                                          FormulaRule(formula=['NOT(ISERROR(SEARCH("' + genotype + '",K2)))'],
                                                      stopIfTrue=True, fill=conditions[genotype]))
        wb.active = ws
        writer.save()  # Save xlsx.
        if self._multi_export:
            print(Message(' Added sheet ' + sheet).timestamp(machine='Export'))
        else:
            print(Message(' ' + os.path.split(self.xlsx_file)[1]).timestamp(machine='Export'))
            self._last_file = self.xlsx_file
            os.startfile(self.xlsx_file)
            self.xlsx_file = None
